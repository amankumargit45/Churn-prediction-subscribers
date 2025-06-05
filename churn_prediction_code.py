import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.tree import DecisionTreeClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import classification_report, confusion_matrix
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# Step 1: Simulate Dataset
np.random.seed(42)
n = 1000

df = pd.DataFrame({
    'subscription_length': np.random.randint(1, 24, n),
    'last_login_days_ago': np.random.randint(0, 60, n),
    'articles_read_per_week': np.random.randint(0, 50, n),
    'engagement_score': np.round(np.random.rand(n), 2),
    'support_tickets_raised': np.random.randint(0, 5, n),
})

# Step 2: Define churn logic
df['churn'] = ((df['last_login_days_ago'] > 30) & 
               (df['engagement_score'] < 0.3) |
               (df['articles_read_per_week'] < 5)).astype(int)

# Step 3: Train/Test Split
X = df.drop('churn', axis=1)
y = df['churn']
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Step 4: Train Models
log_model = LogisticRegression()
tree_model = DecisionTreeClassifier(max_depth=4)

log_model.fit(X_train, y_train)
tree_model.fit(X_train, y_train)

# Step 5: Add Predictions to DataFrame
df['prediction'] = tree_model.predict(X)

# Step 6: Export to Excel
excel_path = "subscriber_churn_dashboard_final.xlsx"
df.to_excel(excel_path, index=False)

# Step 7: Load Excel for Visualization
wb = openpyxl.load_workbook(excel_path)
ws = wb.active
summary_ws = wb.create_sheet(title="Churn Summary")

# Step 8: Pivot - Subscription Length vs Churn
pivot_data = df.groupby(['subscription_length', 'prediction']).size().unstack(fill_value=0)
pivot_data.reset_index(inplace=True)

for row in dataframe_to_rows(pivot_data, index=False, header=True):
    summary_ws.append(row)

# Step 9: Bar Chart for Subscription Length vs Churn
bar_chart = BarChart()
bar_chart.title = "Churn by Subscription Length"
bar_chart.y_axis.title = "Count"
bar_chart.x_axis.title = "Subscription Length"

data = Reference(summary_ws, min_col=2, min_row=1, max_col=3, max_row=summary_ws.max_row)
cats = Reference(summary_ws, min_col=1, min_row=2, max_row=summary_ws.max_row)
bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(cats)
bar_chart.height = 10
bar_chart.width = 20

summary_ws.add_chart(bar_chart, "E2")

# Step 10: Pie Chart for Overall Churn vs Active
churn_counts = df['prediction'].value_counts().sort_index()
summary_ws["H1"] = "Status"
summary_ws["I1"] = "Count"
summary_ws["H2"] = "Active"
summary_ws["H3"] = "Churned"
summary_ws["I2"] = churn_counts.get(0, 0)
summary_ws["I3"] = churn_counts.get(1, 0)

pie = PieChart()
labels = Reference(summary_ws, min_col=8, min_row=2, max_row=3)
data = Reference(summary_ws, min_col=9, min_row=1, max_row=3)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = "Overall Churn Distribution"
pie.height = 8
pie.width = 8

summary_ws.add_chart(pie, "E20")

# Step 11: Add Engagement Bucket and Pivot Table
df['engagement_bucket'] = pd.cut(df['engagement_score'], bins=[0, 0.3, 0.7, 1.0], labels=["Low", "Medium", "High"])
engagement_pivot = df.groupby(['engagement_bucket', 'prediction']).size().unstack(fill_value=0)
engagement_pivot.reset_index(inplace=True)

start_row = summary_ws.max_row + 3
for row in dataframe_to_rows(engagement_pivot, index=False, header=True):
    summary_ws.append([""] * 7 + row)  # Add offset to align columns

# Step 12: Save Final Excel File
wb.save(excel_path)
print(f"Excel dashboard saved as: {excel_path}")